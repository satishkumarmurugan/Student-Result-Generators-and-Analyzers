import os
from werkzeug.utils import secure_filename
import openpyxl
import pandas as pd
from flask import Flask, flash, request, redirect, url_for, render_template, session, send_file
from flask_login import UserMixin, login_user, LoginManager, login_required, logout_user, current_user
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField
from wtforms.validators import DataRequired, InputRequired, Length, ValidationError
from flask_sqlalchemy import SQLAlchemy
import bcrypt
import glob
from openpyxl.styles import Alignment, Font, Border, Side
from process import read_excel_file, convert_to_numeric, find_top_scorers, create_output_dfs, save_to_excel

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
db = SQLAlchemy(app)
app.secret_key = 'secret_key'

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True)
    password = db.Column(db.String(100))

    def __init__(self,email,password,name):
        self.name = name
        self.email = email
        self.password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    def check_password(self,password):
        return bcrypt.checkpw(password.encode('utf-8'),self.password.encode('utf-8'))

with app.app_context():
    db.create_all()


@app.route('/')
def index():
    return render_template('login.html')

@app.route('/register',methods=['GET','POST'])
def register():
    if request.method == 'POST':
        # handle request
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            return redirect('/register')
        new_user = User(name=name,email=email,password=password)
        db.session.add(new_user)
        db.session.commit()
        return redirect('/login')



    return render_template('register.html')

@app.route('/login',methods=['GET','POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        user = User.query.filter_by(email=email).first()
        
        if user and user.check_password(password):
            session['email'] = user.email
            return render_template('index.html')
        else:
            return render_template('login.html',error='Invalid user')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('email',None)
    return redirect('/login')



# Define a global DataFrame to store the loaded Excel data
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def external(filename):
 wb = openpyxl.load_workbook(filename)
 sheet = wb.active

# Create a new workbook for the extracted columns
 new_wb = openpyxl.Workbook()
 new_sheet = new_wb.active

# Extract columns 1, 2, and 3
 new_sheet.append(["Name", "Subject1","Subject2","Subject3","Subject4","Subject5"])
 for row in sheet.iter_rows(min_row=6):
     cell1 = row[7]
     cell2 = row[11]
     cell3 = row[20]
     cell4 = row[29]
     cell5 = row[38]
     cell6 = row[47]
     new_sheet.append([cell1.value, cell2.value, cell3.value, cell4.value, cell5.value, cell6.value])

 # Save the new workbook
 new_filenameexternal ="Theory Grades.xlsx"
 new_wb.save(new_filenameexternal)
 return new_filenameexternal

  

def internal(filename):
 wb = openpyxl.load_workbook(filename)
 sheet = wb.active

# Create a new workbook for the extracted columns
 new_wb = openpyxl.Workbook()
 new_sheet = new_wb.active

# Extract columns 1, 2, and 3
 new_sheet.append(["Name", "Subject1","Subject2","Subject3","Subject4","Subject5"])
 for row in sheet.iter_rows(min_row=6):
     cell1 = row[7]
     cell2 = row[13]
     cell3 = row[22]
     cell4 = row[31]
     cell5 = row[40]
     cell6 = row[49]
     new_sheet.append([cell1.value, cell2.value, cell3.value, cell4.value, cell5.value, cell6.value])
 
  # Save the new workbook
 new_filenameinternal= "Internal Grades.xlsx"
 new_wb.save(new_filenameinternal)
 return new_filenameinternal


@app.route('/report', methods=['POST', 'GET'])
def upload_file_for_report():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'

        file = request.files['file']

        if file.filename == '':
            return 'No selected file'
        if file:
                inter = internal(file)
                exter = external(file)
                df = pd.read_excel(inter)
                df1 = pd.read_excel(exter)
                df3 = pd.read_excel(exter)
                filenames = [inter,exter]

            # Create a list of Workbook objects
                workbooks = []
                for filename in filenames:
                    workbooks.append(openpyxl.load_workbook(filename))
                

            # Create a new workbook for the extracted columns

                # Count the number of "F" grades for each student
                df['F_grades'] = df.apply(lambda row: (row == 'F').sum(), axis=1)

                num_appeared_students = len(df)

                num_fail_students = len(df[df['F_grades'] > 0])

                num_pass_students = len(df[df['F_grades'] == 0])

                # Count the number of students who got at least one "F" grade
                

                num_f_students = len(df[df['F_grades'] == 1])

                # Count the number of students who got two "F" grades
                num_two_f_students = len(df[df['F_grades'] == 2])

                # Count the number of students who got 3 "F" grades
                num_three_f_students = len(df[df['F_grades'] == 3])

                # Count the number of students who got two "F" grades
                num_four_f_students = len(df[df['F_grades'] == 4])
                # Count the number of students who got two "F" grades
                num_five_f_students = len(df[df['F_grades'] == 5])
                # Count the number of students who got two "F" grades
                num_six_f_students = len(df[df['F_grades'] == 6])

                # Count the number of "F" grades for each student
                df1['F_grades'] = df1.apply(lambda row: (row == 'F').sum(), axis=1)

                num_appeared_students1 = len(df1)

                num_fail_students1 = len(df1[df1['F_grades'] > 0])

                num_pass_students1 = len(df1[df1['F_grades'] == 0])
                # Count the number of students who got at least one "F" grade
                num_f_students1 = len(df1[df1['F_grades'] == 1])

                # Count the number of students who got two "F" grades
                num_two_f_students1 = len(df1[df1['F_grades'] == 2])

                # Count the number of students who got 3 "F" grades
                num_three_f_students1 = len(df1[df1['F_grades'] == 3])

                # Count the number of students who got two "F" grades
                num_four_f_students1 = len(df1[df1['F_grades'] == 4])
                # Count the number of students who got two "F" grades
                num_five_f_students1 = len(df1[df1['F_grades'] == 5])
                # Count the number of students who got two "F" grades
                num_six_f_students1 = len(df1[df1['F_grades'] == 6])
            
                
                subjects = ["Subject1","Subject2","Subject3","Subject4","Subject5"]
                grades = ['O', 'A', 'B', 'C', 'D', 'E', 'P','F']
                results = []
                
                for subject in subjects:
                    df3[subject].replace(['O', 'A', 'B', 'C'], 'O to C', inplace=True)
                    df3[subject].replace(['E', 'P'], 'E to P', inplace=True)
                    df3['Pass/Fail'] = df3[subject].apply(lambda x: 'Pass' if x in grades else ('Fail' if x == 'F' else 'Absent'))
                    # Create a dictionary to store the grade counts for the current subject
                    grades_count = {'O to C': 0, 'D': 0, 'E to P': 0, 'F': 0, 'Absent': 0}
                    
                    # Count the number of each grade and total students for the current subject
                    for grade in grades_count.keys():
                        grades_count[grade] = len(df3[df3[subject] == grade])
                    
                    # Calculate total pass and total appeared for the current subject
                    grades_count['Total Students Appeared'] = len(df3)
                    total_pass = grades_count['O to C'] + grades_count['D'] + grades_count['E to P']
                    total_appeared = grades_count['Total Students Appeared']
                    
                    # Calculate pass percentage
                    pass_percentage = round((total_pass / total_appeared) * 100,2)
                    
                    # Add the calculated values to the dictionary
                    grades_count['Total Pass'] = total_pass
                    grades_count['Appeared'] = total_appeared
                    grades_count['Per Pass'] = pass_percentage
                    results.append(grades_count)

                    

                output_df = pd.DataFrame(results, index=subjects, columns=['O to C', 'D', 'E to P', 'F', 'Absent', 'Appeared', 'Total Pass', 'Per Pass'])
                
                output_file = 'Report.xlsx'
                output_df.index.name = 'SUBJECT'

                # Save output to Excel file
                
                output_file = 'Report.xlsx'

                if output_file:
                    # Create a new workbook with xlsxwriter
                    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                        output_df.to_excel(writer, sheet_name='Sheet1', startrow=7, startcol=1, header=True)

                        # Get the worksheet from the writer
                        worksheet = writer.sheets['Sheet1']

                        # Write heading to the first row
                        heading_format = writer.book.add_format({'bold': True, 'font_size': 36})
                        heading_format1 = writer.book.add_format({'bold': True,'italic':True})
                        bold_format = writer.book.add_format({'bold': True,'border': True})
                        a = request.form.get("sem")
                        b = request.form.get("year")
                        c = request.form.get("half")
                        d = int(request.form.get("batch"))
                        f = request.form.get("div")
                        if b == 'FE':
                            e = 'FIRST'
                        elif b == 'SE':
                            e = 'SECOND'
                        elif b == 'TE':
                            e = 'THIRD'
                        elif b == 'BE':
                            e = 'FINAL'
                        else:
                            e = ''
                        border_format = writer.book.add_format({'border': True})
                        worksheet.write(1, 2, 'Result Analysis -SEM ' + a, heading_format)
                        worksheet.write(2, 3, 'DEPARTMENT OF COMPUTER ENGINEERING')
                        worksheet.write(3, 4, 'Result Analysis '+b+' '+f+'-DIV')
                        worksheet.write(4, 3, e +' YEAR SEM-' + a + ' ' + c+'-ST HALF OF '+ str(d)+ ' BATCH RESULT')
                        worksheet.write(24, 2, 'RESULT ANALYSIS COORDINATOR',heading_format1 )
                        worksheet.write(24, 7, 'HOD',heading_format1 )
                        worksheet.write(24, 9, 'PRINCIPAL',heading_format1 )
                        
                        start_row, start_col = 7, 1
                        end_row, end_col = start_row + len(output_df), start_col + len(output_df.columns)
                        worksheet.conditional_format(start_row, start_col, end_row, end_col, {'type': 'no_blanks', 'format': border_format})

                        
                        
                        worksheet.write(17, 0, "External",border_format)
                        worksheet.write(18, 0, "Internal",border_format)
                        

                        # Write the values for each column
                        worksheet.write(16, 4, '1 KT',border_format)
                        worksheet.write(17, 4, num_f_students1,border_format)
                        worksheet.write(18, 4, num_f_students,border_format)
                        

                        # Write other values similarly
                        worksheet.write(16, 5, '2 KT',border_format)
                        worksheet.write(17, 5, num_two_f_students1,border_format)
                        worksheet.write(18, 5, num_two_f_students,border_format)
                        

                        worksheet.write(16, 6, '3 KT',border_format)
                        worksheet.write(17, 6, num_three_f_students1,border_format)
                        worksheet.write(18, 6, num_three_f_students,border_format)
                        

                        worksheet.write(16, 7, '4 KT',border_format)
                        worksheet.write(17, 7, num_four_f_students1,border_format)
                        worksheet.write(18, 7, num_four_f_students,border_format)
                        

                        worksheet.write(16, 8, '5 KT',border_format)
                        worksheet.write(17, 8, num_five_f_students1,border_format)
                        worksheet.write(18, 8, num_five_f_students,border_format)
                        

                        worksheet.write(16, 9, '6 KT',border_format)
                        worksheet.write(17, 9, num_six_f_students1,border_format)
                        worksheet.write(18, 9, num_six_f_students,border_format)

                        worksheet.write(16, 1, 'APPEARED',border_format)
                        worksheet.write(17, 1, num_appeared_students1,border_format)
                        worksheet.write(18, 1, num_appeared_students,border_format)

                        worksheet.write(16, 2, 'FAIL',border_format)
                        worksheet.write(17, 2, num_fail_students1,border_format)
                        worksheet.write(18, 2, num_fail_students,border_format)

                        worksheet.write(16, 3, 'PASS',border_format)
                        worksheet.write(17, 3, num_pass_students1,border_format)
                        worksheet.write(18, 3, num_pass_students,border_format)

                        s1 = request.form.get("s1")
                        s2 = request.form.get("s2")
                        s3 = request.form.get("s3")
                        s4 = request.form.get("s4")
                        s5 = request.form.get("s5")
                        worksheet.write(8, 1, s1,border_format)
                        worksheet.write(9, 1, s2,border_format)
                        worksheet.write(10, 1, s3,border_format)
                        worksheet.write(11, 1, s4,border_format)
                        worksheet.write(12, 1, s5,border_format)

                        si1 = request.form.get("si1")
                        si2 = request.form.get("si2")
                        si3 = request.form.get("si3")
                        si4 = request.form.get("si4")
                        si5 = request.form.get("si5")
                        worksheet.write(7, 0, "FACULTY",bold_format)
                        worksheet.write(8, 0, si1,border_format)
                        worksheet.write(9, 0, si2,border_format)
                        worksheet.write(10, 0, si3,border_format)
                        worksheet.write(11, 0, si4,border_format)
                        worksheet.write(12, 0, si5,border_format)


                        b1=d-1;
                        x=str(b1) +" - " + str(d) + " BATCH NUMBER OF STUDENTS"
                        worksheet.write(15, 4, x ,heading_format1)
                        y = round((num_pass_students1 / num_appeared_students1) * 100, 2)
                        worksheet.write(19, 2, '% ' + x +' =' + str(y),heading_format1)
                        
                    flash('Sucessfull generated the output in report.xlsx')            
    return render_template('report.html')


def merge_and_extract(files, start_index=1, sheet_name='Sheet1'):
    # Check if there are any files to merge
    if not files:
        print("No files found in the upload.")
        return

    # Initialize an empty DataFrame to store the merged data
    merged_data = pd.read_excel(files[0], sheet_name=sheet_name, skiprows=start_index)

    # Iterate through each file and merge its contents with the existing DataFrame
    for file in files[1:]:
        df = pd.read_excel(file, sheet_name=sheet_name, skiprows=start_index)
        merged_data = pd.concat([merged_data, df], axis=1)

    # Ask the user to save the merged file
    output_file = "Merged_Output.xlsx"

    # Write the merged data to a new Excel file starting from the third row
    merged_data.to_excel(output_file, index=False, sheet_name=sheet_name, header=None, startrow=start_index-1)
    print(f"Merged data written to {output_file}")


    # Extract and format columns from the merged file

def extract_and_format(file_path,A,B,D,E):
    try:
        # Read the Excel file using pandas
        df = pd.read_excel(file_path)

        # Extract columns using openpyxl
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Create a new workbook for the extracted columns
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active

        for _ in range(6):
            new_sheet.append([""] * 14)

        new_sheet.cell(1, 3, 'Result Analysis - SEM' +  A).font = Font(size=14)
        new_sheet.cell(2, 3, 'DEPARTMENT OF '+ B)
        new_sheet.cell(3, 3, 'RESULT ANALYSIS')
        new_sheet.cell(4, 3,  D + ' (SEM '+ A+ ') ' + E)

        new_sheet.cell(6,1,'NAME OF SUBJECT')
        new_sheet.cell(6,2,'O to C')
        new_sheet.cell(6,3,'D')
        new_sheet.cell(6,4,'E to P')
        new_sheet.cell(6,5,'Total')
        new_sheet.cell(6,6,'NAME OF SUBJECT')
        new_sheet.cell(6,7,'O to C')
        new_sheet.cell(6,8,'D')
        new_sheet.cell(6,9,'E to P')
        new_sheet.cell(6,10,'Total')
        new_sheet.cell(6,11,'A')
        new_sheet.cell(6,12,'B')
        new_sheet.cell(6,13,'C')
        new_sheet.cell(6,14,'T')

        for row in sheet.iter_rows(min_row=7, max_row=10):
            cell1  = row[1]
            cell2  = row[2]
            cell3  = row[3]
            cell4  = row[4]
            cell5  = row[8]
            cell6  = row[11]
            cell7  = row[12]
            cell8  = row[13]
            cell9  = row[14]
            cell10 = row[18]

            cell100 = row[7]
            cell200 = row[17]

            # Check if cell8.value is not zero before performing division
            if cell8.value != 0:
                cell21 = round(float(cell2.value / cell100.value) * 100, 2)
                cell31 = round(float(cell3.value / cell100.value) * 100, 2)
                cell41 = round(float(cell4.value / cell100.value) * 100, 2)
                cell51 = round(float(cell5.value / cell100.value) * 100, 2)
                cell71 = round(float(cell7.value / cell200.value) * 100, 2)
                cell81 = round(float(cell8.value / cell200.value) * 100, 2)
                cell91 = round(float(cell9.value / cell200.value) * 100, 2)
                cell101 = round(float(cell10.value / cell200.value) * 100, 2)
            else:
                # Handle the case where cell8.value is zero (division by zero)
                cell21 = cell31 = cell41 = cell71 = cell81 = cell91 = 0.0

            # Calculate cell11 value
            cell11 = int(cell21) - int(cell71)

            cell12 = int(cell31) - int(cell81)
            cell13 = int(cell41) - int(cell91)
            cell14 = int(cell51) - int(cell101)

            # Append the values to the new sheet
            # Append blank rows before the extracted values
            new_sheet.append([cell1.value, cell21, cell31, cell41, cell51, cell6.value, cell71, cell81, cell91, cell101, cell11, cell12, cell13, cell14])

        for row in sheet.iter_rows(min_row=14, max_row=23):
            extracted_values = [cell.value for cell in row]
            new_sheet.append(extracted_values)

        # Apply borders to the cells in the specified range (row 6 to 14, columns 1 to 14)
        for row_num in range(6, 15):
            if row_num != 11:
                for col_num in range(1, 21):
                    cell_value = new_sheet.cell(row=row_num, column=col_num).value
                    if cell_value is not None:
                        new_sheet.cell(row=row_num, column=col_num).border = Border(left=Side(style='thin'),
                                                                                    right=Side(style='thin'),
                                                                                    top=Side(style='thin'),
                                                                                    bottom=Side(style='thin'))

        # Save the new workbook with a dynamic filename
        new_filename = "Formatted_Output.xlsx"
        new_wb.save(new_filename)

        print("Columns extracted and formatted, and saved to:", new_filename)

    except pd.errors.EmptyDataError:
        print("The selected Excel file is empty.")
    except Exception as e:
        print("An error occurred:", e)


@app.route('/compare', methods=['GET', 'POST'])
def compare():
    if request.method == 'POST':
        if 'files[]' in request.files:
            files = request.files.getlist('files[]')

            if not files:
                return "File selection canceled."
            else:
                merge_and_extract(files)
                return "Files merged and extracted successfully!"

        elif 'file' in request.files:
            file = request.files['file']
            
            if not file:
                return "File selection canceled."
            else:
                a = request.form.get("sem")
                b = request.form.get("dep")
                f = request.form.get("year")
                d = str(request.form.get("batch"))
                if f == 'FE':
                    B = 'FIRST YEAR'
                elif f == 'SE':
                    B = 'SECOND YEAR'
                elif f == 'TE':
                    B = 'THIRD YEAR'
                elif f == 'BE':
                    B = 'FINAL YEAR'
                else:
                    B = ''
                extract_and_format(file,a,b,B,d)
                return "Files Compared successfully!!"

        else:
            return 'No file part'

    return render_template('compare.html')


@app.route('/upload')
def upload():
    return render_template('upload.html')

@app.route('/topper', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return render_template('upload.html', error='No file part')

    file = request.files['file']
    if file.filename == '':
        return render_template('upload.html', error='No selected file')

    file_path = 'uploads/' + file.filename
    file.save(file_path)

    # Process the Excel file
    df = read_excel_file(file_path)
    exam_columns = ['exam3', 'exam6', 'exam9', 'exam12', 'exam15', 'exam18']
    df = convert_to_numeric(df, exam_columns)
    top_scorers = find_top_scorers(df, exam_columns, 5)
    output_dfs = create_output_dfs(top_scorers)

    # Save the output file to the user's desktop
    desktop_dir = os.path.join(os.path.expanduser('~'), 'Desktop')
    if not os.path.exists(desktop_dir):
        os.makedirs(desktop_dir)

    output_file_path = os.path.join(desktop_dir, 'output.xlsx')
    save_to_excel(output_dfs, output_file_path)
    

    return send_file(output_file_path, as_attachment=True)




if __name__ == '__main__':
    app.run(debug=True)
